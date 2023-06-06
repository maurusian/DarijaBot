import pywikibot
from sys import argv
from copy import deepcopy
from params import MONTHS
from datetime import datetime, timezone
from pywikibot.exceptions import LockedPageError
from openpyxl import load_workbook
import re, os

DOC_TMP_1 = "{{توثيق}}"
DOC_TMP_1_bis = "{{/توثيق}}"
DOC_TMP_2 = "{{توتيق}}"
DOC_TMP_3 = "{{documentation}}"
DOC_TMP_3_bis = "{{Documentation}}"
DOC_TMP_4 = "{{شرح}}"

doc_page = "/شرح"

TEMPLATE_NAMESPACE = 10

TASK_LOG_PAGE = "مستخدم:DarijaBot/عطاشة 17: زّيادة ديال شرح لقوالب بشكل ؤطوماتيكي"

LOG_ERROR_COMMENT = "ميساج د ليرور تزاد ف لّوحة"

CASCADE_ERROR_MSG = "لپاج [[{}]] ماتقدرش تبدّل حيت محمية ؤلا حيت غادي تأتر ف پاجات خرين"

LAST_RUN_FILE = "last_run_doc_replace.txt"

DATE_FORMAT = "%Y-%m-%d %H:%M"

UPDATE_TMP_DOC_CALL_MSG = "إصلاح سّمية د لقالب د شّرح"

GENERAL_UPDATE = "إصلاحات عامة"

UNUSED_TEMPLATE_CODE = "{{قالب مامخدمش|{number}}}"
UNUSED_TEMP_PART = "{{قالب مامخدمش"
NOINCLUDE_TAG = "<noinclude>{}</noinclude>"
TOO_MANY_NEXT_LINE = "\n\n\n"
TOO_MANY_NOINCLUDES = "</noinclude><noinclude>"
TOO_MANY_NOINCLUDES_2 = """</noinclude>
<noinclude>"""
TOO_MANY_NOINCLUDES_3 = """</noinclude>

<noinclude>"""

MAINTENANCE_CAT = "تصنيف:قوالب ديال صيانة"

WRONG_DOC_HEADER = """<noinclude>{{طرجامة}}
{{صفحات فرعية د شرح}}</noinclude>"""

CORRECT_DOC_HEADER = """{{طرجامة}}
<noinclude>{{صفحات فرعية د شرح}}</noinclude>"""

DELETION_TAG = "<noinclude>{{حدف|1=صيانة}}</noinclude>"

DELETION_TAG_PART = "{{حدف"

LOG_ERROR_COMMENT = "ميساج د ليرور تزاد ف لّوحة"

CASCADE_ERROR_MSG = "لپاج [[{}]] ماتقدرش تبدّل حيت محمية ؤلا حيت غادي تأتر ف پاجات خرين"

TASK_LOG_PAGE = "مستخدم:DarijaBot/عطاشة 17: زّيادة ديال شرح لقوالب بشكل ؤطوماتيكي"

TRANSLATION_FILE = "to_translate.xlsx"

def get_translations():
    wb = load_workbook(TRANSLATION_FILE)

    sheet = wb["Sheet2"]

    translation_dict = {}

    for i in range(2,sheet.max_row+1):
        translation_dict[sheet["A"+str(i)].value] = sheet["C"+str(i)].value

    return translation_dict

def log_error(LOG_PAGE_TITLE,log_message,save_log_message,site):
    log_page = pywikibot.Page(site, LOG_PAGE_TITLE)
    if log_page.text != '':
        log_page.text += '\n* '+log_message
    else:
        log_page.text = '* '+log_message

    log_page.save(save_log_message)

def get_last_run_datetime():
    if not os.path.exists(LAST_RUN_FILE):
        with open(LAST_RUN_FILE,'w') as f:
            return None

    with open(LAST_RUN_FILE,'r') as f:
        datetime_str = f.read().strip()

    return datetime.strptime(datetime_str,DATE_FORMAT)

def write_run_time():
    with open(LAST_RUN_FILE,'w') as f:
        f.write(pywikibot.Timestamp.now().strftime(DATE_FORMAT))
    

def get_time_string():
    raw_time = pywikibot.Timestamp.now(tz=timezone.utc)
    #utc_time = datetime.now(tz=timezone.utc)
    raw_time_parts = str(raw_time).split('T')
    date_parts = raw_time_parts[0].split('-')
    return " "+raw_time_parts[1][:-4]+"، "+date_parts[2]+" "+MONTHS[int(date_parts[1])-1]["ary_name"]+" "+date_parts[0]+" (UTC)"

if __name__ == '__main__':
    print(len(argv))
    if len(argv)>2:
        local_args = argv[2:]
    else:
        local_args = None

    #local_args = 0 
    if local_args is not None:
        site = pywikibot.Site()
        if local_args[-1] == '-l':
            last_run_time = get_last_run_datetime()
            print(last_run_time)
            print("running for last changed pages")
            #load last changed
            last_changes = site.recentchanges(namespaces=[TEMPLATE_NAMESPACE], reverse=True, top_only=True, start=last_run_time, redirect = False)
            #create page pool
            #NEXT: check other potential last_change types

            pool = [pywikibot.Page(site, item['title']) for item in last_changes]

        else:
            print("running for all templates")
            #load all pages on the article namespace, default option
            pool = site.allpages(namespace=TEMPLATE_NAMESPACE) #, filterredir=False)

            #redirect_pool = site.allpages(namespace=TEMPLATE_NAMESPACE,filterredir=True)
        
        pool_size = len(list(deepcopy(pool)))
        print('Pool size: '+str(pool_size))
        i = 1
        translation_dict = get_translations()
        for page in pool:
            
            print('*********'+str(i)+'/'+str(pool_size))
            
            text = page.text.replace(DOC_TMP_1,DOC_TMP_4).replace(DOC_TMP_2,DOC_TMP_4).replace(DOC_TMP_3,DOC_TMP_4).replace(DOC_TMP_3_bis,DOC_TMP_4)
            text = text.replace(TOO_MANY_NEXT_LINE,"\n")
            text = text.replace(TOO_MANY_NOINCLUDES,"")
            text = text.replace(TOO_MANY_NOINCLUDES_2,"")
            text = text.replace(TOO_MANY_NOINCLUDES_3,"")

            if doc_page in page.title():
                if page.isRedirectPage():
                    text = DELETION_TAG + text
                else:
                    
                    tmp_page = pywikibot.Page(site,page.title()[:-4])
                    if tmp_page.text == '' or tmp_page.isRedirectPage():
                        if not DELETION_TAG_PART in text:
                            text = DELETION_TAG + text
                    else:
                        text = text.replace(WRONG_DOC_HEADER,CORRECT_DOC_HEADER)

                    for key,value in translation_dict.items():
                        pattern = r"==\s*"+key+r"\s*=="
                        text = re.sub(pattern,"== "+value+" ==",text)
                        #print("testing title substitution")
                        #print(text)
                        #input()

            """
            if "/" not in page.title():

                backlinks = [p.title() for p in page.backlinks()]
                categories = [c.title() for c in page.categories()]

                if MAINTENANCE_CAT not in categories:
                    if len(backlinks) == 0:
                        text = NOINCLUDE_TAG.format(UNUSED_TEMPLATE_CODE.replace("{number}","1"))+text
                    elif len(backlinks) == 1 and backlinks[0]==page.title()+doc_page:
                        text = NOINCLUDE_TAG.format(UNUSED_TEMPLATE_CODE.replace("{number}","2"))+text
                    elif len(backlinks) > 1:
                        all_irrelevant_backlink = True
                        for link in page.backlinks():
                            if UNUSED_TEMP_PART not in link.text: 
                                all_irrelevant_backlink = False
                            elif link.title()==page.title()+doc_page:
                                all_irrelevant_backlink = False
                        if all_irrelevant_backlink:
                            text = NOINCLUDE_TAG.format(UNUSED_TEMPLATE_CODE.replace("{number}","3"))+text
            
            #break
            """
            if text != page.text:
                page.text = text
                try:
                    page.save(GENERAL_UPDATE)
                except LockedPageError:
                    log_message = CASCADE_ERROR_MSG.format(page.title())
                    save_log_message = LOG_ERROR_COMMENT
                    log_error(TASK_LOG_PAGE,log_message,save_log_message,site)
                #break
            i+=1
        write_run_time()

from pgvbotLib import *
from params import *
import pywikibot
from copy import deepcopy
import re
from openpyxl import load_workbook

entries = './export/page_redirects.xlsx'

site = pywikibot.Site()

#load content of xlsx into dict
wb = load_workbook(entries)
sheet = wb.active
numbers_dict = {}
months_dict = {}
for i in range(2,sheet.max_row+1):
    try:
        numbers_dict[int(sheet['A'+str(i)].value)] = [x.strip() for x in sheet['B'+str(i)].value.strip().split('..')]
    except ValueError:
        months_dict[sheet['A'+str(i)].value] = [x.strip() for x in sheet['B'+str(i)].value.strip().split('..')]

#print(len(numbers_dict))
#print(len(months_dict))
'''
for key,value in numbers_dict.items():
    print(str(key)+': '+str(value))

for key,value in months_dict.items():
    print(key+': '+str(value))

'''
#add flat entries
for key, value in numbers_dict.items():
    for entry in value:
        page = pywikibot.Page(site,entry)
        if page.text == '':
            page.text = MOVE_TEXT+' [['+str(key)+']]\n\n'+REDIRECT_PAGE_CAT_CODE
            MESSAGE = "تحويلة تقادات"
            save_page(page,MESSAGE)
print('Numbers done')
for key, value in months_dict.items():
    for entry in value:
        page = pywikibot.Page(site,entry)
        if page.text == '':
            page.text = MOVE_TEXT+' [['+key+']]\n\n'+REDIRECT_PAGE_CAT_CODE
            MESSAGE = "تحويلة تقادات"
            save_page(page,MESSAGE)
print('Months are done')
#add combined entries
daymonths = []
for month in MONTHS:
    for i in range(month['day_count']):
        daymonths.append((i+1,month['ary_name'],str(i+1)+' '+month['ary_name']))
#print(len(daymonths))

print('starting treatment of Daymonths (first pass) 0/366')
i = 1
for daymonth in daymonths:
    for day_entry in numbers_dict[daymonth[0]]:
        page = pywikibot.Page(site,day_entry+' '+daymonth[1])
        if page.text == '':
            page.text = MOVE_TEXT+' [['+daymonth[2]+']]\n\n'+REDIRECT_PAGE_CAT_CODE
            MESSAGE = "تحويلة تقادات"
            save_page(page,MESSAGE)
    print(str(i)+'/366')
    i+=1

print('starting treatment of Daymonths (second pass) 0/366')
i = 1
for daymonth in daymonths:
    for day_entry in numbers_dict[daymonth[0]]:
        for month_entry in months_dict[daymonth[1]]:
            page = pywikibot.Page(site,day_entry+' '+month_entry)
            if page.text == '':
                page.text = MOVE_TEXT+' [['+daymonth[2]+']]\n\n'+REDIRECT_PAGE_CAT_CODE
                MESSAGE = "تحويلة تقادات"
                save_page(page,MESSAGE)

    print(str(i)+'/366')
    i+=1

        

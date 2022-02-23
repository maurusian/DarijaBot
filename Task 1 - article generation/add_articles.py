from pgvbotLib import *
import pywikibot
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from datetime import datetime
from months import MONTHS
from urllib.parse import unquote
import re


PERSON_INFO_TAG         = ""
XML_MODEL               = "model.xml"
CAT_MODEL               = "categories.xml"

DATA_SOURCE             = "Moroccan_politicians.xlsx"

F_MARKER1               = "ة"
F_MARKER2               = "ت"
F_MARKER3               = "ات"

PM_MARKER               = "ين" #plural masculine
PF_MARKER               = "ات" #plural feminine

F                       = "female"

INDEPENDENT             = "لامنتامي"

#to be extracted from herka.xml in the future
HERKA_MAIN_NOMUNCLATURE = "سياسي"
HERKA_MAIN_CITIZENSHIP  = "مغريبي"
BORN                    = "مزيود"
DIED                    = "مات"


FULL_DATE_FORMAT        = "%Y-%m-%d"
MONTH_YEAR_DATE_FORMAT  = "%Y-%m"

PAST_TENSE_BE           = "كان"

BELONG_MALE_FORM        = "كاينتامي"
BELONG_FEMALE_FORM      = "كاتنتامي"
GRADUATED               = "تخرّج"

TRANSLATION_TABLE       = {"and":"ؤ"
                          ,"or":"ؤلا"}

COMMON_FOOTER           = """== عيون لكلام ==
{{عيون}}

{{ضبط مخازني}}
{{زريعة شخصيات د لمغريب}}"""

COMMON_HEADER           = """{{واساخ خاصو يتراجع|arwiki={arwiki}|enwiki={enwiki}|frwiki={frwiki}}}
{{معلومات شخص}}"""

DRAFT_ADDED_MESSAGE     = "لواساخ تزاد"

DRAFT_NS                = "واساخ"

def wrap(word, wrapper):

    if wrapper == "[[":
        return "[["+word+"]]"
    elif wrapper == "{":
        return "{"+word+"}"
    elif wrapper == "{{":
        return "{{"+word+"}}"

    elif wrapper == " ":
        return " "+word+" "

    elif wrapper == "'''":
        return " '''"+word+"''' "
    
    return word

def get_object_from_cell_value(cell_value):
    #print(cell_value)
    #print(type(cell_value))
    #default value
    lam = lambda x:x
    if type(cell_value) == str:
        if ".." in cell_value:
            lam = lambda x:[y.strip() for y in x.split("..")] if x is not None and x.strip() is not None else ""
        elif "https://" in cell_value or "http://" in cell_value:
            lam = lambda x:unquote(x.split("/")[-1].replace("_"," ")) if x is not None and x.strip() is not None else ""
        
    
    elif type(cell_value) == datetime:
        lam = lambda x:x if x is not None else ""
        

    return lam(cell_value)

def get_month_ary_name(month_number):
    lam = lambda x:MONTHS[int(x)-1]["ary_name"] if x is not None else ""

    return lam(month_number)

def load_wkdt_politician_values():
    wb = load_workbook(DATA_SOURCE)
    sheet = wb.active

    politicians_data = []

    for i in range(1,sheet.max_row):
        politician        = {}

        #extract raw values from xlsx file before treatment
        wkdt_link         = sheet["A"+str(i+1)].value
        ary_name          = sheet["D"+str(i+1)].value
        gender            = sheet["F"+str(i+1)].value
        party             = sheet["G"+str(i+1)].value
        occupations       = sheet["I"+str(i+1)].value
        positions         = sheet["K"+str(i+1)].value
        schools           = sheet["L"+str(i+1)].value
        citizenships      = sheet["M"+str(i+1)].value
        birth_date_string = sheet["O"+str(i+1)].value
        death_date_string = sheet["Q"+str(i+1)].value
        birth_place       = sheet["P"+str(i+1)].value
        death_place       = sheet["R"+str(i+1)].value
        custom_text       = sheet["T"+str(i+1)].value
        enwiki            = sheet["U"+str(i+1)].value
        arwiki            = sheet["V"+str(i+1)].value
        frwiki            = sheet["W"+str(i+1)].value

        #treatment of values
        if ary_name is not None and ary_name.strip() is not None:
            #Wikidata code
            
            politician["wkdt_code"] = get_object_from_cell_value(wkdt_link)
            #(lambda wkdt_link:wkdt_link.split("/")[-1] if wkdt_link is not None and wkdt_link.strip() is not None else "")(wkdt_link)

            #name in Moroccan Darija
            politician["person_full_name"] = ary_name

            #gender markers
            
            politician["f_marker"] = (lambda gender:F_MARKER1 if gender == F else "")(gender)
            politician["f_marker2"] = (lambda gender:F_MARKER2 if gender == F else "")(gender)
            politician["f_marker3"] = (lambda gender:F_MARKER3 if gender == F else "")(gender)

            politician["born"] = BORN+politician["f_marker"]
            politician["died"] = DIED+politician["f_marker2"]
            politician["graduated"] = GRADUATED+politician["f_marker3"]

            if gender == F:
                politician["belong_gender_form"] = BELONG_FEMALE_FORM
            else:
                politician["belong_gender_form"] = BELONG_MALE_FORM

            #party
            
            politician["party_name"] = (lambda party:party.strip() if party is not None and party.strip() is not None and party != INDEPENDENT else "")(party)
            politician["independent"] = (lambda party:party+politician["f_marker"] if party == INDEPENDENT else "")(party)
                

            #occupations
            
            politician["main_occupation"] = HERKA_MAIN_NOMUNCLATURE+politician["f_marker"]
            occ = get_object_from_cell_value(occupations)
            if type(occ) == list or occ is None:
                politician["other_occupations"] = occ
            else:
                politician["other_occupations"] = [occ]
            if politician["other_occupations"] is not None and politician["other_occupations"] != "":
                for i in range(len(politician["other_occupations"])):
                    politician["other_occupations"][i]+=politician["f_marker"]
            #(lambda occupations:[occupation.strip() for occupation in occupations.split("..")] if occupations is not None and occupations.strip() is not None else "")(occupations)

            #positions
            pos = get_object_from_cell_value(positions)
            if type(pos) == list or pos is None:
                politician["positions"] = pos
            else:
                politician["positions"] = [pos]
            #(lambda positions:[position.strip() for position in positions.split("..")] if positions is not None and positions.strip() is not None else "")(positions)

            #schools
            schools = get_object_from_cell_value(schools)
            if type(schools) == list or schools is None:
                politician["schools"] = schools
            else:
                politician["schools"] = [schools]
            #(lambda schools:[school.strip() for school in schools.split("..")] if schools is not None and schools.strip() is not None else "")(schools)

            #citizenships
            
            politician["main_citizenship"] = HERKA_MAIN_CITIZENSHIP+politician["f_marker"]
            citizenships = get_object_from_cell_value(citizenships)
            if type(citizenships) == list:
                citizenships = citizenships[0]
            
            politician["other_citizenship"] = (lambda citizenships:citizenships+politician["f_marker"] if citizenships is not None and len(citizenships)>0 else "")(citizenships)
            #(lambda citizenships:[citizenship.strip() for citizenship in citizenships.split("..")] if citizenships is not None and citizenships.strip() is not None else "")

            #birthdate
            
            birth_date = get_object_from_cell_value(birth_date_string)
            #(lambda birth_date_string:birth_date_string if birth_date_string is not None else "")(birth_date_string)
            if birth_date is not None:
                politician["birth_day_number"] = birth_date.strftime("%d").lstrip("0")
                month_number = birth_date.strftime("%m")
                politician["birth_month_string_ary"] = get_month_ary_name(month_number)
                politician["birth_year"] = birth_date.strftime("%Y")
            else:
                politician["birth_day_number"] = ""
                politician["birth_month_string_ary"] = ""
                politician["birth_year"] = ""

            #birthplace
            politician["birth_location"] = birth_place

            #deathdate
            
            death_date = get_object_from_cell_value(death_date_string)
            #print("death_date_string type: "+str(type(death_date_string)))
            #print(death_date_string)
            #(lambda death_date_string:death_date_string if death_date_string is not None else "")(death_date_string)
            if death_date is not None:
                politician["death_day_number"] = death_date.strftime("%d").lstrip("0")
                #print(politician["death_day_number"])
                #print(death_date)
                #print(type(death_date))
                month_number = death_date.strftime("%m")
                politician["death_month_string_ary"] = get_month_ary_name(month_number)
                politician["death_year"] = death_date.strftime("%Y")
            else:
                politician["death_day_number"] = ""
                politician["death_month_string_ary"] = ""
                politician["death_year"] = ""
                

            #deathplace
            politician["death_location"] = death_place
            
            #past tense indicator 1 and 2
            if death_date is not None or death_place is not None:
                politician["past_tense_indicator"] = PAST_TENSE_BE+politician["f_marker2"]

            else:
                politician["past_tense_indicator"] = ""


            politician["past_tense_indicator2"] = PAST_TENSE_BE+politician["f_marker2"]


            if (birth_date is not None or birth_place is not None) and (death_date is not None or death_place is not None):
                politician["comma"] = "،"
            else:
                politician["comma"] = ""

            #custom text
            politician["custom_text"] = custom_text

            #enwiki
            
            politician["enwiki"] = str(get_object_from_cell_value(enwiki) or '')
            #(lambda enwiki:unquote(enwiki) if enwiki is not None and enwiki.strip() is not None else "")

            #arwiki
            
            politician["arwiki"] = str(get_object_from_cell_value(arwiki) or '')
            #(lambda arwiki:unquote(arwiki) if arwiki is not None and arwiki.strip() is not None else "")

            #frwiki
            
            politician["frwiki"] = str(get_object_from_cell_value(frwiki) or '')
            #(lambda frwiki:unquote(frwiki) if frwiki is not None and frwiki.strip() is not None else "")
            
            politicians_data.append(politician)
    
    return politicians_data

def fix_text_errors(text):

    text = re.sub(' +',' ',text)
    text = re.sub(' \.','.',text)
    text = re.sub(' ،','،',text)
    text = re.sub(' \)',')',text)
    text = re.sub('\( ','(',text)
    #replace("  "," ").replace(" .",".").replace(" ،","،").replace(" )",")").replace("( ","(")

    return text

def isEmptyObject(obj):
    #print("test if object is empty")
    #print(type(obj))
    if type(obj) == str:
        
        if len(obj.strip()) == 0:
            #print("object is really empty")
            return True
        
    elif type(obj) == list:
        if len(obj) == 0:
            #print("object is really empty")
            return True
        
    else:
        if obj is None:
            #print("object is really empty")
            return True

    return False

def process_sentence(sentence,values):
    params            = sentence.findall("param")
    sub_sentences     = sentence.findall("sentence")
    sentence_str      = sentence.attrib["value"]
    #print(sentence.attrib["name"])
    conditionals      = sentence.attrib["conditional"].split(",")
    #print(conditionals)
    conditional_flags = [True for x in conditionals]
    replaceable       = None
    replaced          = None
    i=0 #conditional flags counter
    for param in params:
        param_name = param.attrib["name"]
        #print(param_name)
        replaceable = None
        if param.attrib["type"] == "var":
            if param_name not in conditionals or not isEmptyObject(values[param_name]):
                replaceable   = wrap(param_name,"{")
                if not isEmptyObject(values[param_name]):
                    #print(values[param_name])
                    if "wrapper" in param.attrib.keys():
                        wrapper = param.attrib["wrapper"]
                        #print("wrapper: "+wrapper)
                        replaced      = wrap(values[param_name].strip(),wrapper)
                        #print("replaced: "+replaced)
                    else:
                        replaced      = wrap(values[param_name]," ")
                else:
                    replaced  = ""
            else:
                #print("setting flag to False for param "+param_name)
                conditional_flags[i] = False
                i+=1
            
        elif param.attrib["type"] == "added_list":
            if param_name not in conditionals or values[param_name] is not None:
                replaceable   = wrap(param_name,"{")
                ary_separator = TRANSLATION_TABLE[param.attrib["separator"]]
                if not isEmptyObject(values[param_name]):
                    #print(values[param_name])
                    replaced  = wrap(ary_separator+SPACE+(SPACE+ary_separator+SPACE).join(values[param_name])," ")
                else:
                    replaced  = ""
            else:
                conditional_flags[i] = False
                i+=1
        elif param.attrib["type"] == "list":
            if param_name not in conditionals or values[param_name] is not None:
                replaceable   = wrap(param_name,"{")
                ary_separator = TRANSLATION_TABLE[param.attrib["separator"]]
                if not isEmptyObject(values[param_name]):
                    #print(values[param_name])
                    replaced  = wrap((SPACE+ary_separator+SPACE).join(values[param_name])," ")
                else:
                    replaced  = ""
            else:
                conditional_flags[i] = False
                i+=1
        elif param.attrib["type"] == "sub":
            for sub_sentence in sub_sentences:
                if sub_sentence.attrib["name"] == param_name:
                    replaceable   = wrap(param_name,"{")
                    subtext = process_sentence(sub_sentence,values)
                    if param_name not in conditionals or subtext.strip() != "":
                        replaced = wrap(subtext," ")
                    else:
                        conditional_flags[i] = False
                        i+=1
                    break

        if replaceable is not None and replaced is not None:
            sentence_str  = sentence_str.replace(replaceable,replaced)
        
    sentence_str = fix_text_errors(sentence_str)

    #re.sub(' +',' ',sentence_str.replace("  "," ").replace(" .",".").replace(" ،","،").replace(" )",")").replace("( ","("))

    #print(sentence.attrib["value"])
    #print("Conditional value (test15) "+str(any(conditional_flags))+" ("+str(conditional_flags)+") for sentence "+sentence.attrib["name"]+" with phrase "+sentence.attrib["value"])
    if not any(conditional_flags):
        sentence_str = ""

    return sentence_str


def process_articles(rec):
    #article = [] #list of paragraphs
    tree = ET.parse(XML_MODEL)
    root = tree.getroot()

    #print(dir(root))

    for art in root.findall("article"):
        if len(art)>0:
            article = ""
            for par in art:
                paragraph = ""
                for sentence in par:
                    paragraph+=process_sentence(sentence,rec).strip()+'\n'
                article+=paragraph.strip()+'\n\n'
            return article.strip()
                
        else:
            return ""

def load_categories(data):
    tree = ET.parse(CAT_MODEL)
    root = tree.getroot()
    cats = []
    for category in root.findall("category"):
        valid_flag = True
        cat = category.attrib["value"]
        for param in category:
            param_name = param.attrib["name"]
            if isEmptyObject(data[param_name]):
                valid_flag = False
                break
            cat = cat.replace(wrap(param_name,"{"),data[param_name])
        if valid_flag:
            cats.append(wrap(cat,'[['))
    
    return '\n'.join(cats)


def get_header(data):
    return COMMON_HEADER.replace("{arwiki}",data["arwiki"]).replace("{enwiki}",data["enwiki"]).replace("{frwiki}",data["frwiki"])


def export_page_to_draft(title,text):
    site = pywikibot.Site()
    page = pywikibot.Page(site,title)
    page.text = text
    page.save(DRAFT_ADDED_MESSAGE)

if __name__ == "__main__":

    politicians_data = load_wkdt_politician_values()
    counter = 1
    total   = len(politicians_data)
    for politician in politicians_data:
        print(str(counter)+"/"+str(total))
        sections = []
        sections.append(get_header(politician))
        sections.append(process_articles(politician))
    
        sections.append(COMMON_FOOTER)
        
        sections.append(load_categories(politician))
        #break
        text = '\n\n'.join([section.strip() for section in sections])
        export_page_to_draft(DRAFT_NS+':'+politician["person_full_name"],text)

        """
        with open(politician["person_full_name"]+".txt",'w',encoding="utf-8") as f:
            f.write(text)
        """
#site = pywikibot.Site()



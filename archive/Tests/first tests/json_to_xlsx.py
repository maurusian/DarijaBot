from openpyxl import Workbook, load_workbook

TRANSLATION_TABLE = {'politician':'سياسي','artist':'فنان','lawyer':'محامي',
                     'physician':'طبيب','diplomat':'دبلوماسي','engineer':'مهندس','writer':'كاتب',
                     'Arabic':'لعربية','French':'لفرانساوية','Moroccan Arabic':'داريجة','Modern Standard Arabic':'لعربية',
                     'Dutch':'لهولاندية','Berber languages':'لأمازيغية','Spanish':'صبليونية','Hebrew':'لعبرية','English':'نڭليزية',
                     'Rabat':'رباط','Tangier':'طنجة','Fez':'فاس','Casablanca':'كازا','Morocco':'لمغريب'}

def translate(label):
    try:
        return TRANSLATION_TABLE[label]
    except:
        return label

def clean(v):
   
    return ' .. '.join([translate(el) for el in set([x.strip() for x in v.split('..')])])
    



ORD_A = ord('A')
filename = "./data/Moroccan_politicians.json"
EXPORT = "./export/Moroccan_politicians.xlsx"
HEADERS = ["politician", "politicianLabel", "nativeName", "sexLabel", "parties",
           "occupations","positions","education","citizenships","languages",
           "dateOfBirth","placeOfBirthLabel","dateOfDeath","placeOfDeathLabel",
           "mannerOfDeathLabel","articleEN","articleAR","articleFR"]																


with open(filename,'r',encoding='utf8') as f:
    crypto_dict_list = eval(f.read())


"""
headers = set()
for crypto_dict in crypto_dict_list:
    temp = set(crypto_dict.keys())
    headers = headers | temp
"""

#wb = Workbook()
wb = load_workbook(EXPORT)

sheet = wb.active


for i in range(len(HEADERS)):
    sheet[chr(ORD_A+i)+'1'] = HEADERS[i]
for i in range(len(crypto_dict_list)):
    for key, value in crypto_dict_list[i].items():
        if 'date' in key:
            sheet[chr(ORD_A+HEADERS.index(key))+str(i+2)] = value.split('T')[0]
        else:
            sheet[chr(ORD_A+HEADERS.index(key))+str(i+2)] = clean(value)
        
    




wb.save(EXPORT)

from lib.params import *
from openpyxl import load_workbook
import pywikibot, sys


DATA = "./data/Cryptos filtered.xlsx"
SHEET_NAME = "translated"

wb = load_workbook(DATA)

sheet = wb[SHEET_NAME]

def set_args_to_none(*args):
    for arg in args:
        arg = None

    return args

def get_header(source):
    if source is None:
        return ADD_SOURCE_TAG+NEXT_LINE+CRYPTO_INFO_TAG
    else:
        return CRYPTO_INFO_TAG

def get_logo_code(name,logo_url):
    return PICTURE_TEMPLATE.format(logo_url,name)

def get_or_bold_names(name_value):
    return (SPACE+OR+SPACE).join([NAME_TEXT_TEMPLATE.format(name.strip()) for name in name_value.split('..')])

def get_and_names(name_value):
    name_parts = [name.strip() for name in name_value.split('..')]

    if len(name_parts)>1:
        return (COMMA+SPACE).join(name_parts[:-1])+SPACE+AND+SPACE+name_parts[-1]
    else:
        return name_parts[0]

def get_additional_info(original_name,ticker,unit_symbol):
    
    text_parts = []

    if original_name is not None:
        ORIGIN_TEXT = ORIGIN_TEXT_TEMPLATE.format(get_or_bold_names(original_name))
        text_parts.append(ORIGIN_TEXT)
    
    if unit_symbol is not None:
        SYMBOL_TEXT = SYMBOL_TEXT_TEMPLATE.format(get_or_bold_names(unit_symbol))
        text_parts.append(SYMBOL_TEXT)
    
    if ticker is not None:
        TICKER_TEXT = TICKER_TEXT_TEMPLATE.format(get_or_bold_names(ticker))
        text_parts.append(TICKER_TEXT)

    return "("+(COMMA+SPACE).join(text_parts)+")"

def convert_date(date):
    
    dash_count = date.count(DASH)
    if dash_count == 2:
        date_parts = date.split(DASH)
        year_value = date_parts[0]
        month_value = date_parts[1]
        day_value = date_parts[2]
        return day_value+SPACE+CONVERSION_DICT[month_value]+SPACE+year_value

    elif dash_count == 1:
        date_parts = date.split(DASH)
        year_value = date_parts[0]
        month_value = date_parts[1]
        
        return SPACE+CONVERSION_DICT[month_value]+SPACE+year_value
    
    return date

def get_invention_text(creation_date,creators):

    text_parts = []
    if creation_date is not None:
        text_parts.append(INVENTION_DATE_TEXT_TEMPLATE.format(creation_date))
    if creators is not None:
        text_parts.append(INVENTOR_TEXT_TEMPLATE.format(get_and_names(creators)))

    return SPACE.join(text_parts)

def get_definition_text(definition,end_date):
    if definition is None:
        if end_date is None:
            definition = IS_PRESENT+SPACE+DEFAULT_DEFINITION_TEXT
        else:
            definition = IS_PAST+SPACE+DEFAULT_DEFINITION_TEXT

    return definition

def get_intro_text(name_text,additional_info_text,definition,invention_text):
    intro_text = name_text

    if additional_info_text != "":
        intro_text += SPACE+additional_info_text

    intro_text += SPACE+definition

    if invention_text != "":
        intro_text += SPACE+invention_text

    return intro_text+DOT

def get_tech_text(name,technologies):
    return TECHNOLOGY_TEXT_TEMPLATE.format(name,get_and_names(technologies))

def get_purpose_text(purpose):
    return PURPOSE_TEXT_TEMPLATE.format(get_and_names(purpose))

def get_prog_text(languages):
    return PROG_TEXT_TEMPLATE.format(get_and_names(languages))

def get_copyright_text(copyrights):
    return COPYRIGHT_TEXT_TEMPLATE.format(get_and_names(copyrights))

def get_inspiration_text(inspiration):
    return INSPIRATION_TEXT_TEMPLATE.format(get_and_names(inspiration))

def get_details1_text(tech_text,purpose_text,prog_text,copyright_text,inspiration_text,exchanges_text):
    details1_text_parts = []

    if tech_text != "":
        details1_text_parts.append(tech_text)

    if purpose_text != "":
        details1_text_parts.append(purpose_text)

    if prog_text != "":
        details1_text_parts.append(prog_text)

    if copyright_text != "":
        details1_text_parts.append(copyright_text)

    if inspiration_text != "":
        details1_text_parts.append(inspiration_text)


    details1_text = get_and_names('..'.join(details1_text_parts))+DOT

    if exchanges_text != "":
        details1_text += SPACE+exchanges_text+DOT

    return details1_text

def get_end_project_text(end_date):
    
    return END_PROJECT_TEXT_TEMPLATE.format(end_date)

def get_etymology_text(name,etymology):
    return ETYMOLOGY_TEXT_TEMPLATE.format(name,etymology)

def get_details2_text(end_text,etymology_text):
    details2_text_parts = []
    if end_text != "":
        details2_text_parts.append(end_text)

    if etymology_text != "":
        details2_text_parts.append(etymology_text)

    return get_and_names('..'.join(details2_text_parts))+DOT

def get_exchanges_text(exchanges):
    return EXCHANGE_TEXT_TEMPLATE.format(get_and_names(exchanges))
    
def get_source_text(source,original_name,name):
    return SOURCE_TEXT_TEMPLATE.format(original_name,source,name)


def get_full_text(header_text,logo_text,intro_text,details1_text,details2_text,custom_text,source_text,footer):
    full_text = ""
    if logo_text != "":
        full_text+=logo_text+NEXT_LINE

    
    full_text += header_text


    full_text += NEXT_LINE+intro_text

    if details1_text != "":
        full_text+= NEXT_LINE*2+details1_text

    if details2_text != "":
        full_text+= NEXT_LINE*2+details2_text

    if custom_text is not None:
        full_text+= NEXT_LINE*2+custom_text

    if source_text != "":
        full_text+= source_text

    full_text += footer

    return full_text

def export_to_file(filepath,text):
    try:
        with open(filepath,'w',encoding="utf8") as f:
            f.write(text)
        print(filepath+' exported')
    except:
        print(sys.exc_info())

def save_to_wiki(site,title,text):
    page = pywikibot.Page(site, title)
    if page.text == "":
        page.text = text
        page.save(SAVE_MESSAGE.format(title))
    else:
        print("Page "+title+" already exists!")


#main program
#should not run without args from CMD
#print(sheet.max_row)
site = pywikibot.Site()
for i in range(2,sheet.max_row):

    #print(i)

    #original_name, name_value, name, definition, unit_symbol, ticker, creation_date, end_date, creators, purpose, technologies, copyrights, languages, exchanges, inspiration, etymology, logo_url
    
    original_name = sheet['B'+str(i)].value
    name_value = sheet['C'+str(i)].value
    if name_value is None:
        break

    name_value = name_value.strip()
    name = [name.strip() for name in name_value.split('..')][0]
    definition = sheet['D'+str(i)].value
    unit_symbol = sheet['E'+str(i)].value
    ticker = sheet['F'+str(i)].value
    creation_date = sheet['G'+str(i)].value
    if creation_date is not None:
        #print(type(creation_date))
        if type(creation_date).__name__ == 'int':
            creation_date = str(creation_date)
        else:
            creation_date = creation_date.strftime('%Y-%m-%d')
    end_date = sheet['H'+str(i)].value

    if end_date is not None:
        if type(end_date).__name__ == 'int':
            end_date = str(end_date)
        else:
            end_date = end_date.strftime('%Y-%m-%d')

    creators = sheet['I'+str(i)].value
    purpose = sheet['J'+str(i)].value
    technologies = sheet['K'+str(i)].value
    copyrights = sheet['L'+str(i)].value
    languages = sheet['M'+str(i)].value
    exchanges = sheet['O'+str(i)].value
    inspiration = sheet['P'+str(i)].value
    etymology = sheet['Q'+str(i)].value
    logo_url = sheet['R'+str(i)].value
    source = sheet['T'+str(i)].value
    custom_text = sheet['U'+str(i)].value

    #1- header
    header_text = get_header(source)

    #2- logo image
    logo_code_text = ""
    if logo_url is not None:
        logo_code_text = get_logo_code(name,logo_url)

    #3- intro
    name_text = get_or_bold_names(name_value)

    additional_info_text = ""
    if original_name is not None or ticker is not None or unit_symbol is not None:
        additional_info_text = get_additional_info(original_name,ticker,unit_symbol)



    invention_text = ""
    if creation_date is not None or creators is not None:
        invention_text = get_invention_text(convert_date(creation_date),creators)

    definition_text = get_definition_text(definition,end_date)


    intro_text = get_intro_text(name_text,additional_info_text,definition_text,invention_text)

    #4- details 1
    tech_text = ""
    if technologies is not None:
        tech_text = get_tech_text(name,technologies)

    purpose_text = ""
    if purpose is not None:
        purpose_text = get_purpose_text(purpose)
        #print(purpose_text)

    prog_text = ""
    if languages is not None:
        prog_text = get_prog_text(languages)
        #print(prog_text)

    copyright_text = ""
    if copyrights is not None:
        copyright_text = get_copyright_text(copyrights)
        #print(copyright_text)

    inspiration_text = ""
    if inspiration is not None:
        inspiration_text = get_inspiration_text(inspiration)
        #print(inspiration_text)

    exchanges_text = ""
    if exchanges is not None:
        exchanges_text = get_exchanges_text(exchanges)

    details1_text = ""
    if tech_text != "" or purpose_text != "" or prog_text != "" or copyright_text != "" or inspiration_text != "" or exchanges_text != "":
        details1_text = get_details1_text(tech_text,purpose_text,prog_text,copyright_text,inspiration_text,exchanges_text)

    #print(details1_text)

    #5- details 2
    
    end_text = ""
    if end_date is not None:
        end_text = get_end_project_text(end_date)
        #print(end_text)

    etymology_text = ""
    if etymology is not None:
        etymology_text = get_etymology_text(name,etymology)
        #print(etymology_text)

    details2_text = ""
    if etymology_text != "" or end_text != "":
        details2_text = get_details2_text(end_text,etymology_text)

    #print(details2_text)

    #6- custom text
    if custom_text is not None:
        custom_text = custom_text.strip()+DOT

    #7- source
    source_text = ""
    if source is not None:
        source_text = get_source_text(source,original_name,name)

    full_text = get_full_text(header_text,logo_code_text,intro_text,details1_text,details2_text,custom_text,source_text,FOOTER)

    #print(full_text)

    filepath = name+'.txt'

    export_to_file(filepath,full_text)

    
    if i>24:
        break
    save_to_wiki(site,name,full_text)

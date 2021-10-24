from openpyxl import Workbook, load_workbook

    

ORD_A = ord('A')


filename = "./data/query.json"
EXPORT = "./data/export.xlsx"
HEADERS = ["person", "dateofbirth", "yearofbirth", "dateofdeath", "yearofdeath"]																


with open(filename,'r',encoding='utf8') as f:
    dict_list = eval(f.read())


"""
headers = set()
for crypto_dict in crypto_dict_list:
    temp = set(crypto_dict.keys())
    headers = headers | temp
"""

wb = Workbook()
#wb = load_workbook(EXPORT)

sheet = wb.active




for i in range(len(HEADERS)):
    sheet[chr(ORD_A+i)+'1'] = HEADERS[i]

j=0
for i in range(len(dict_list)):
    valid_rec = False
    if 'personLabel' in dict_list[i].keys() and dict_list[i]['personLabel'].strip() != '':
        if ('dateOfBirth' in dict_list[i].keys() and dict_list[i]['dateOfBirth'] != '') or ('dateOfDeath' in dict_list[i].keys() and dict_list[i]['dateOfDeath'] != ''):
            
            if ('dateOfBirth' in dict_list[i].keys() and dict_list[i]['dateOfBirth'] != ''):
                fulldob = dict_list[i]['dateOfBirth'].split('T')[0]
                if len(fulldob.split('-')) == 3:
                    valid_rec = True
            if ('dateOfDeath' in dict_list[i].keys() and dict_list[i]['dateOfDeath'] != ''):
                fulldod = dict_list[i]['dateOfDeath'].split('T')[0]
                if len(fulldod.split('-')) == 3:
                    valid_rec = True
            if valid_rec:
                sheet['A'+str(i+2)] = dict_list[i]['personLabel']
                if ('dateOfBirth' in dict_list[i].keys() and dict_list[i]['dateOfBirth'] != ''):
                    sheet['B'+str(j+2)] = fulldob[4:].replace('-','')
                    sheet['C'+str(j+2)] = fulldob[:4]
                
                if ('dateOfDeath' in dict_list[i].keys() and dict_list[i]['dateOfDeath'] != ''):
                    sheet['D'+str(j+2)] = fulldod[4:].replace('-','')
                    sheet['E'+str(j+2)] = fulldod[:4]
                j+=1
    else:
        print(i)
                


wb.save(EXPORT)

import pywikibot
from openpyxl import load_workbook

def check_emails_in_multiple_projects(excel_path):
    # Load the Excel workbook and select the first sheet
    workbook = load_workbook(filename=excel_path)
    sheet = workbook.active

    # Skip the header row, start from the second row
    for row in range(2, sheet.max_row + 1):
        username_cell = sheet.cell(row=row, column=1)
        username = username_cell.value

        if username is not None:
            # Project names can be found in columns B, G, L (2, 7, 12)
            for project_col, output_col in [(2, 4), (7, 9), (12, 14)]:
                project_cell = sheet.cell(row=row, column=project_col)
                project = project_cell.value

                if project is not None and project in ['enwiki', 'arwiki', 'frwiki', 'itwiki']:
                    site = pywikibot.Site(project[:-4], project[-4:]+'pedia')
                    user = pywikibot.User(site, username)

                    if user.isRegistered():
                        if user.has_email:
                            status = 'Has email'
                        else:
                            status = 'No email'
                    else:
                        status = 'Not registered'

                    # Write the status to the output column of the same row
                    sheet.cell(row=row, column=output_col, value=status)

    # Save the changes to the Excel file
    workbook.save(excel_path)

# Usage example
check_emails_in_multiple_projects("Moroccan Wikimedians.xlsx")

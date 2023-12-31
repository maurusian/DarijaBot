import requests
from openpyxl import load_workbook, Workbook

read_excel_filename = "user_contribs.xlsx"
target_excel_filename = "user_contribs_target.xlsx"

def get_user_contributions(username):
    # Define the URL for the API
    url = "https://meta.wikimedia.org/w/api.php"

    # Define parameters for querying the user's global contributions
    params = {
        "action": "query",
        "format": "json",
        "meta": "globaluserinfo",
        "guiuser": username,
        "guiprop": "merged"
    }

    # Query the user's global contributions
    response = requests.get(url, params=params)
    response_json = response.json()

    print(response_json.keys())

    # Query the user's global information
    response = requests.get(url, params=params)
    response_json = response.json()

    # Check if the response contains the expected data
    if "query" in response_json and "globaluserinfo" in response_json["query"] and "merged" in response_json["query"]["globaluserinfo"]:
        merged_accounts = response_json["query"]["globaluserinfo"]["merged"]
        
        # Filter out accounts with 0 edits and sort by edit count
        non_zero_edits = filter(lambda x: x["editcount"] > 0, merged_accounts)
        sorted_accounts = sorted(non_zero_edits, key=lambda x: x["editcount"], reverse=True)

        # Extract project and editcount
        contributions = [(account["wiki"], account["editcount"]) for account in sorted_accounts]

        return contributions
    else:
        # Print error message if the response is not what we expect
        print("Error:", response_json.get("error", "Unexpected API response"))
        return None

def write_to_excel(max_row_num):
    load_wb = load_workbook(read_excel_filename)
    load_sheet = load_wb.active
    #print(load_sheet.max_row)
    
    for i in range(2,max_row_num+1):
        username = load_sheet["B"+str(i)].value
        print(username)
        user_contribs = get_user_contributions(username)

        if user_contribs is not None:
        
            j = 0
            for project, count in user_contribs[:5]:
                load_sheet[chr(67+j)+str(i)] = project
                print(chr(68+j)+str(i),count)
                load_sheet[chr(68+j)+str(i)] = count
                j+=2

    load_wb.save(target_excel_filename)
        
write_to_excel(546)
# Example usage:
username = "Ideophagous"  # Replace with the desired username
contributions = get_user_contributions(username)
for project, count in contributions:
    print(f"{project}: {count} contributions")
